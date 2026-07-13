"""
Script untuk membaca SEMUA sheet dari data_dip_real.xlsx
dan generate DipSeeder.php yang lengkap dan terurut dengan benar.
Kategori disesuaikan dengan kode sistem (lowercase dengan tanda hubung).
Tabel informasi_dikecualikans juga diisi otomatis untuk data DIK.
"""
import openpyxl
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

# Mapping sheet ke kategori DIP sistem
SHEET_KATEGORI = {
    'DIP Setiap Saat':  'informasi-setiap-saat',
    'DIP Sertamerta':   'informasi-serta-merta',
    'DIP Berkala':      'informasi-berkala',
    'DIK':              'informasi-dikecualikan',
}

wb = openpyxl.load_workbook('data_dip_real.xlsx', data_only=True)

def clean(val):
    if val is None:
        return ''
    s = str(val).strip()
    s = s.replace('\n', ' ').replace('\r', ' ')
    # Remove multiple spaces
    s = re.sub(r'\s+', ' ', s)
    return s

def php_escape(s):
    return s.replace("\\", "\\\\").replace("'", "\\'")

def parse_no(val):
    if val is None:
        return 999999.0
    try:
        s = str(val).strip()
        # Find numeric parts
        match = re.search(r'\d+(\.\d+)?', s)
        if match:
            return float(match.group())
    except:
        pass
    return 999999.0

all_records = []
dik_records = []

for sheet_name in wb.sheetnames:
    kategori = SHEET_KATEGORI.get(sheet_name, sheet_name)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=False))
    
    # Find header row index
    header_idx = None
    for i, row in enumerate(rows[:20]):
        non_none = [c.value for c in row if c.value is not None]
        # Header row has "INFORMASI" and "RINGKASAN" keywords
        row_str = ' '.join(str(c) for c in non_none).upper()
        if 'INFORMASI' in row_str and 'RINGKASAN' in row_str and len(non_none) >= 5:
            header_idx = i
            break
    
    if header_idx is None:
        print(f"WARNING: Could not find header in sheet {sheet_name}")
        continue
    
    headers = [clean(c.value).upper().replace('\n', ' ') for c in rows[header_idx]]
    
    # Map column positions
    col_no       = next((i for i, h in enumerate(headers) if 'NO' in h and len(h.strip()) <= 8), 0)
    col_info     = next((i for i, h in enumerate(headers) if 'INFORMASI' in h and 'RINGKASAN' not in h and 'JENIS' not in h), 1)
    col_ring     = next((i for i, h in enumerate(headers) if 'RINGKASAN' in h), 2)
    col_pejabat  = next((i for i, h in enumerate(headers) if 'PEJABAT' in h or 'MENGUASAI' in h), 3)
    col_penerbit = next((i for i, h in enumerate(headers) if 'PENERBIT' in h), 4)
    col_bentuk   = next((i for i, h in enumerate(headers) if 'BENTUK' in h), 5)
    col_tempat   = next((i for i, h in enumerate(headers) if 'TEMPAT' in h or 'WAKTU PEMBUATAN' in h), 6)
    col_jangka   = next((i for i, h in enumerate(headers) if 'JANGKA' in h or 'RETENSI' in h), 7)
    
    col_asli = next((i for i, h in enumerate(headers) if 'LINK DOKUMEN ASLI' in h), -1)
    col_prev = next((i for i, h in enumerate(headers) if 'PREVIEW' in h or 'UPLOAD' in h), -1)
    col_sens = next((i for i, h in enumerate(headers) if 'SENSOR' in h), -1)
    
    sheet_records = []
    
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if row is None:
            continue
        
        def get_col(idx):
            if idx != -1 and idx < len(row):
                return clean(row[idx].value)
            return ''
            
        def get_col_link(idx):
            if idx != -1 and idx < len(row):
                cell = row[idx]
                if cell.hyperlink and cell.hyperlink.target:
                    return str(cell.hyperlink.target).strip()
                val = str(cell.value).strip() if cell.value is not None else ''
                if val.startswith('http') or 'drive.google.com' in val:
                    return val
            return ''
        
        no_val = get_col(col_no)
        info   = get_col(col_info)
        
        # Skip empty rows / total rows / header-like rows
        if not info or info.upper() in ('INFORMASI', '') :
            continue
        if not no_val and not info:
            continue
        # Skip rows that look like headers or totals
        if 'JUMLAH' in info.upper() and len(info) < 20:
            continue
        
        # Parse links
        link_asli = get_col_link(col_asli)
        link_prev = get_col_link(col_prev)
        link_sens = get_col_link(col_sens)
        
        # Prioritize: Sensor (safe) -> Preview/Upload -> Asli
        best_link = ''
        for link in [link_sens, link_prev, link_asli]:
            link_clean = link.strip()
            if link_clean and (link_clean.startswith('http') or 'drive.google.com' in link_clean):
                best_link = link_clean
                break
        
        # Skip if no document link is found (requested by client)
        if not best_link:
            continue
            
        # Parse waktu pembuatan dari kolom tempat/waktu
        waktu_raw = get_col(col_tempat)
        # Try to extract year from the tempat+waktu string
        year_match = re.search(r'20\d{2}', waktu_raw)
        year = year_match.group() if year_match else '2025'
        
        # Build waktu_pembuatan (e.g. "Tegal, 2025")
        waktu_pembuatan = waktu_raw if waktu_raw else f'Tegal, {year}'
        
        # created_at: use year to sort newest first
        created_at = f'{year}-01-01 00:00:00'
        
        record = {
            'judul_informasi': php_escape(info),
            'kategori': kategori,
            'isi_informasi': php_escape(get_col(col_ring)),
            'pejabat_penguasa': php_escape(get_col(col_pejabat)),
            'penanggung_jawab': php_escape(get_col(col_penerbit)),
            'penerbit_informasi': php_escape(get_col(col_penerbit)),
            'bentuk_informasi': php_escape(get_col(col_bentuk)),
            'tempat_pembuatan': 'Tegal',
            'waktu_pembuatan': php_escape(waktu_pembuatan),
            'jangka_waktu': php_escape(get_col(col_jangka)),
            'file_informasi': php_escape(best_link),
            'aktif': 1,
            'created_at': created_at,
            'updated_at': created_at,
            'sort_year': int(year),
            'sort_no': parse_no(row[col_no].value),
            'orig_index': i
        }
        sheet_records.append(record)
        
        # Jika kategori ini dikecualikan, simpan salinan untuk tabel informasi_dikecualikans
        if kategori == 'informasi-dikecualikan':
            dik_rec = {
                'judul': php_escape(info),
                'deskripsi': php_escape(get_col(col_ring)),
                'tanggal': f'{year}-01-01',
                'jangka_waktu': php_escape(get_col(col_jangka)),
                'penanggung_jawab': php_escape(get_col(col_penerbit)),
                'file_path': php_escape(best_link),
                'file_name': 'Dokumen',
                'file_size': '-',
                'file_type': 'Google Drive / PDF',
                'aktif': 1,
                'is_blurred': 0,
                'bisa_download': 1,
                'created_at': created_at,
                'updated_at': created_at,
                'sort_year': int(year),
                'sort_no': parse_no(row[col_no].value),
                'orig_index': i
            }
            dik_records.append(dik_rec)
    
    # Sort sheet records
    sheet_records.sort(key=lambda x: x['orig_index'])
    sheet_records.sort(key=lambda x: x['sort_no'])
    sheet_records.sort(key=lambda x: x['sort_year'], reverse=True)
    
    print(f"Sheet [{sheet_name}] -> {len(sheet_records)} records, kategori: {kategori}")
    all_records.extend(sheet_records)

# Sort dik records
dik_records.sort(key=lambda x: x['orig_index'])
dik_records.sort(key=lambda x: x['sort_no'])
dik_records.sort(key=lambda x: x['sort_year'], reverse=True)

print(f"\nTotal records daftar_informasis: {len(all_records)}")
print(f"Total records informasi_dikecualikans: {len(dik_records)}")

# Generate PHP seeder
php_lines = []
php_lines.append("<?php\n")
php_lines.append("namespace Database\\Seeders;\n\n")
php_lines.append("use Illuminate\\Database\\Seeder;\n")
php_lines.append("use Illuminate\\Support\\Facades\\DB;\n\n")
php_lines.append("class DipSeeder extends Seeder\n{\n")
php_lines.append("    public function run()\n    {\n")
php_lines.append("        // 1. Seed daftar_informasis\n")
php_lines.append("        DB::table('daftar_informasis')->truncate();\n\n")
php_lines.append("        $data = [\n")

for rec in all_records:
    php_lines.append("        [\n")
    for k, v in rec.items():
        if k in ('sort_year', 'sort_no', 'orig_index'):
            continue
        if k in ('aktif',):
            php_lines.append(f"            '{k}' => {v},\n")
        else:
            php_lines.append(f"            '{k}' => '{v}',\n")
    php_lines.append("        ],\n")

php_lines.append("        ];\n\n")
php_lines.append("        foreach (array_chunk($data, 50) as $chunk) {\n")
php_lines.append("            DB::table('daftar_informasis')->insert($chunk);\n")
php_lines.append("        }\n\n")

# Seed informasi_dikecualikans
php_lines.append("        // 2. Seed informasi_dikecualikans\n")
php_lines.append("        DB::table('informasi_dikecualikans')->truncate();\n\n")
php_lines.append("        $data_dik = [\n")

for rec in dik_records:
    php_lines.append("        [\n")
    for k, v in rec.items():
        if k in ('sort_year', 'sort_no', 'orig_index'):
            continue
        if k in ('aktif', 'is_blurred', 'bisa_download'):
            php_lines.append(f"            '{k}' => {v},\n")
        else:
            php_lines.append(f"            '{k}' => '{v}',\n")
    php_lines.append("        ],\n")

php_lines.append("        ];\n\n")
php_lines.append("        foreach (array_chunk($data_dik, 50) as $chunk) {\n")
php_lines.append("            DB::table('informasi_dikecualikans')->insert($chunk);\n")
php_lines.append("        }\n")

php_lines.append("    }\n}\n")

output = ''.join(php_lines)

with open('database/seeders/DipSeeder.php', 'w', encoding='utf-8') as f:
    f.write(output)

print(f"\nDipSeeder.php ditulis: {len(all_records)} daftar_informasis dan {len(dik_records)} informasi_dikecualikans")
print("SELESAI!")
